#
# Copyright (c) 2026 Masaki Nakai (https://github.com/masakinakai3)
# Released under the MIT license
# https://opensource.org/licenses/mit-license.php
#

"""CSV, XLSX, and JSON export endpoints."""

import io
import json
import re
from datetime import date

from flask import Blueprint, Response, request
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import Allocation, Member, Theme, db

export_bp = Blueprint('export', __name__)


def _rate_fill(rate):
    if rate > 100:
        return PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    if rate == 100:
        return PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    if rate > 60:
        return PatternFill(start_color='B3B3FF', end_color='B3B3FF', fill_type='solid')
    if rate > 30:
        return PatternFill(start_color='E0E0FF', end_color='E0E0FF', fill_type='solid')
    return PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')


def _parse_rate(value):
    if value in (None, ''):
        return None

    text = str(value).strip()
    if not text:
        return None
    if text.endswith('%'):
        text = text[:-1]

    try:
        return int(float(text))
    except ValueError:
        return None


def _configure_header(worksheet, headers):
    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _export_list_layout(workbook, headers, rows):
    worksheet = workbook.active
    worksheet.title = 'Gantt'
    _configure_header(worksheet, headers)

    for row_index, row_data in enumerate(rows, start=2):
        worksheet.append(row_data)
        for column_index, value in enumerate(row_data, start=1):
            if column_index <= 3:
                continue

            rate = _parse_rate(value)
            if rate is None:
                continue

            cell = worksheet.cell(row=row_index, column=column_index)
            cell.fill = _rate_fill(rate)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.value = rate / 100
            cell.number_format = '0%'

    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['C'].width = 15
    for index in range(4, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 12


def _export_gantt_layout(workbook, headers, rows):
    worksheet = workbook.active
    worksheet.title = 'Gantt'
    _configure_header(worksheet, headers)

    group_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    summary_fill = PatternFill(start_color='EAF2F8', end_color='EAF2F8', fill_type='solid')

    for row_index, row_data in enumerate(rows, start=2):
        label = row_data.get('label', '')
        values = row_data.get('values', [])
        row_type = row_data.get('type', 'member')
        worksheet.append([label, *values])

        first_cell = worksheet.cell(row=row_index, column=1)
        first_cell.alignment = Alignment(horizontal='left', vertical='center')

        if row_type == 'group':
            first_cell.fill = group_fill
            first_cell.font = Font(bold=True)
            for column_index in range(2, len(headers) + 1):
                worksheet.cell(row=row_index, column=column_index).fill = group_fill
            continue

        if row_type == 'summary':
            first_cell.fill = summary_fill
            first_cell.font = Font(bold=True)
        else:
            first_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        for column_index, value in enumerate(values, start=2):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(horizontal='center', vertical='center')

            rate = _parse_rate(value)
            if rate is None:
                continue

            cell.value = rate / 100
            cell.number_format = '0%'
            cell.fill = _rate_fill(rate)
            if row_type == 'summary':
                cell.font = Font(bold=True)

    worksheet.freeze_panes = 'B2'
    worksheet.column_dimensions['A'].width = 36
    for index in range(2, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 12


@export_bp.route('/csv', methods=['POST'])
@login_required
def export_csv():
    """Return client-generated CSV as a downloadable file."""
    if request.is_json:
        data = request.get_json()
        csv_content = data.get('content', '')
        filename = data.get('filename', 'gantt_export.csv')
    else:
        csv_content = request.form.get('content', '')
        filename = request.form.get('filename', 'gantt_export.csv')

    filename = re.sub(r'[^\w\-.]', '_', filename)
    if not csv_content:
        return {'error': 'No CSV content provided'}, 400

    response = Response(
        '\ufeff' + csv_content,
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Type': 'text/csv; charset=utf-8',
        },
    )
    return response


@export_bp.route('/xlsx', methods=['POST'])
@login_required
def export_xlsx():
    """Export gantt rows as an XLSX file."""
    data = request.get_json()
    if not data:
        return {'error': 'No data'}, 400

    headers = data.get('headers', [])
    rows = data.get('rows', [])
    filename = re.sub(r'[^\w\-.]', '_', data.get('filename', 'gantt_export.xlsx'))

    workbook = Workbook()
    if data.get('layout') == 'gantt':
        _export_gantt_layout(workbook, headers, rows)
    else:
        _export_list_layout(workbook, headers, rows)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(
        output.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@export_bp.route('/json', methods=['GET'])
@login_required
def export_json():
    """Export all application data as JSON.

    The exported themes include the ``milestones`` array and
    ``dev_complete_month`` so that a full round-trip is possible.
    """
    from models import ThemeMilestone  # local import to avoid circular issues
    themes = [theme.to_dict() for theme in Theme.query.order_by(Theme.theme_id).all()]
    members = [member.to_dict() for member in Member.query.order_by(Member.member_id).all()]
    allocations = [allocation.to_dict() for allocation in Allocation.query.all()]

    theme_members = []
    for theme in Theme.query.all():
        for member in theme.members:
            theme_members.append({'theme_id': theme.theme_id, 'member_id': member.member_id})

    payload = {
        'version': 2,
        'exported_at': db.session.execute(db.select(db.func.datetime('now'))).scalar(),
        'themes': themes,
        'members': members,
        'allocations': allocations,
        'theme_members': theme_members,
    }

    filename = f'manage_backup_{date.today().strftime("%Y%m%d")}.json'
    return Response(
        json.dumps(payload, ensure_ascii=False, indent=2),
        mimetype='application/json; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
